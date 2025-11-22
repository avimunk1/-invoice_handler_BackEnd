from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import ORJSONResponse, FileResponse
from .models import ProcessRequest, ProcessResponse
from .pipeline import process_path, process_path_with_llm
from typing import Dict, Any, List, Optional
from pathlib import Path
from urllib.parse import unquote
from pydantic import BaseModel, Field
from datetime import date
from .config import settings
from .s3_client import get_s3_client
import json
from sqlalchemy import select, insert, update, join
from sqlalchemy.dialects.postgresql import insert as pg_insert
from .database import get_engine, close_engine
from .models_db import suppliers, invoices, customers

app = FastAPI(title="Invoice Handler", default_response_class=ORJSONResponse)

# CORS middleware for frontend
# Allow local development + production domains
allowed_origins = [
	"http://localhost:5173",
	"http://localhost:3000",
	"https://invoice-handler-frontend-2deslnnp5-avimunk1s-projects.vercel.app",  # Vercel production
]

app.add_middleware(
	CORSMiddleware,
	allow_origin_regex=r"https://.*\.vercel\.app|http://localhost:\d+",
	allow_credentials=True,
	allow_methods=["*"],
	allow_headers=["*"],
)


@app.get("/healthz")
async def healthz():
	return {"status": "ok"}


@app.post("/process", response_model=ProcessResponse)
async def process(req: ProcessRequest):
	results, total_files, files_handled = await process_path(
		req.path, req.recursive, req.language_detection, req.starting_point
	)
	return ProcessResponse(
		results=results, 
		errors=[], 
		total_files=total_files,
		files_handled=files_handled,
		vat_rate=settings.vat_rate
	)


@app.post("/process/llm", response_model=ProcessResponse)
async def process_with_llm(req: ProcessRequest):
	"""Process documents using Azure OCR + OpenAI LLM for flexible field extraction."""
	results, total_files, files_handled = await process_path_with_llm(
		req.path, req.recursive, req.language_detection, req.starting_point
	)
	return ProcessResponse(
		results=results, 
		errors=[], 
		total_files=total_files,
		files_handled=files_handled,
		vat_rate=settings.vat_rate
	)


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)) -> Dict[str, Any]:
	"""Upload a file to configured upload directory"""
	from datetime import datetime
	import shutil
	
	# Use configured upload directory
	upload_dir = Path(settings.upload_dir)
	upload_dir.mkdir(parents=True, exist_ok=True)
	
	# Generate unique filename with timestamp
	timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
	safe_filename = file.filename.replace(" ", "_")
	unique_filename = f"{timestamp}_{safe_filename}"
	file_path = upload_dir / unique_filename
	
	# Save file to disk
	try:
		with file_path.open("wb") as buffer:
			shutil.copyfileobj(file.file, buffer)
		
		return {
			"success": True,
			"filename": unique_filename,
			"path": str(file_path),
			"upload_dir": str(upload_dir),
			"original_filename": file.filename
		}
	except Exception as e:
		return {
			"success": False,
			"error": str(e)
		}


@app.post("/upload-and-process", response_model=ProcessResponse)
async def upload_and_process(files: List[UploadFile] = File(...)) -> ProcessResponse:
	"""
	Upload multiple files and process them in one atomic operation.
	1. Clear input directory
	2. Save uploaded files
	3. Process with LLM
	4. Return results
	"""
	from datetime import datetime
	import shutil
	import glob
	
	# Use configured upload directory
	upload_dir = Path(settings.upload_dir)
	upload_dir.mkdir(parents=True, exist_ok=True)
	
	try:
		# Step 1: Save all uploaded files to input directory
		# NOTE: We do NOT clean the directory here because:
		# - Existing files may be from previous uploads that haven't been saved yet
		# - Files only move to 'processed/' after being saved to the database
		# - The user may be adding more files to an existing batch
		saved_files = []
		for file in files:
			timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")  # Include microseconds for uniqueness
			safe_filename = file.filename.replace(" ", "_")
			unique_filename = f"{timestamp}_{safe_filename}"
			file_path = upload_dir / unique_filename
			
			with file_path.open("wb") as buffer:
				shutil.copyfileobj(file.file, buffer)
			
			saved_files.append(str(file_path))
			print(f"[INFO] Saved file: {unique_filename}")
		
		# Step 2: Process only the files we just uploaded
		from .pipeline import process_specific_files_with_llm
		results, total_files, files_handled = await process_specific_files_with_llm(
			file_paths=saved_files,
			language_detection=True
		)
		
		print(f"[INFO] Processed {files_handled} of {total_files} files successfully")
		
		return ProcessResponse(
			results=results,
			errors=[],
			total_files=total_files,
			files_handled=files_handled,
			vat_rate=settings.vat_rate
		)
		
	except Exception as e:
		print(f"[ERROR] Upload and process failed: {str(e)}")
		import traceback
		traceback.print_exc()
		return ProcessResponse(
			results=[],
			errors=[str(e)],
			total_files=0,
			files_handled=0,
			vat_rate=settings.vat_rate
		)


@app.get("/file/view", response_model=None)
async def view_file(path: str = None, filename: str = None):
	"""Get a viewable URL for a file (S3 presigned URL or local file). Auto-converts HEIC to JPEG."""
	# Support both old 'path' parameter and new 'filename' parameter
	if filename:
		# New way: construct path from filename using configured upload_dir
		from .config import settings
		from pathlib import Path
		
		# Try input directory first
		input_path = Path(settings.upload_dir) / filename
		# Also check processed directory
		processed_dir = Path(settings.upload_dir).parent / "processed"
		processed_path = processed_dir / filename
		
		if input_path.exists():
			decoded_path = f"file://{input_path.absolute()}"
		elif processed_path.exists():
			decoded_path = f"file://{processed_path.absolute()}"
		else:
			return {"error": f"File not found: {filename}"}
	elif path:
		# Old way: use provided path
		decoded_path = unquote(path)
	else:
		return {"error": "Either 'path' or 'filename' parameter is required"}
	
	if decoded_path.startswith("s3://"):
		# Generate presigned GET URL for S3
		if not settings.s3_bucket:
			return {"error": "S3 bucket not configured"}
		
		# Parse S3 path
		import re
		m = re.match(r"s3://([^/]+)/(.+)", decoded_path)
		if not m:
			return {"error": "Invalid S3 path"}
		
		bucket = m.group(1)
		key = m.group(2)
		
		s3_client = get_s3_client()
		
		# Check if HEIC/HEIF - need to convert before serving
		if key.lower().endswith(('.heic', '.heif')):
			# Download, convert, and serve as JPEG
			obj = s3_client.get_object(Bucket=bucket, Key=key)
			heic_content = obj["Body"].read()
			
			# Convert to JPEG
			from .pipeline import _convert_heic_to_jpeg
			from io import BytesIO
			jpeg_content = _convert_heic_to_jpeg(heic_content)
			
			# Return JPEG directly
			from fastapi.responses import Response
			return Response(
				content=jpeg_content,
				media_type="image/jpeg",
				headers={
					"Content-Disposition": f'inline; filename="{Path(key).stem}.jpg"'
				}
			)
		else:
			# Generate presigned URL for viewing (1 hour expiry)
			url = s3_client.generate_presigned_url(
				'get_object',
				Params={'Bucket': bucket, 'Key': key},
				ExpiresIn=3600
			)
			return {"url": url}
	
	elif decoded_path.startswith("file://"):
		# Serve local file directly
		local_path = decoded_path[7:]  # Remove file://
		file_path = Path(local_path)
		
		# If file not found, check if it was moved to 'processed' directory
		if not file_path.exists():
			# Try to find in processed directory
			if 'input' in file_path.parts:
				# Replace 'input' with 'processed' in the path
				parts = list(file_path.parts)
				input_idx = parts.index('input')
				parts[input_idx] = 'processed'
				processed_path = Path(*parts)
				if processed_path.exists():
					file_path = processed_path
				else:
					return {"error": "File not found"}
			else:
				return {"error": "File not found"}
		
		# Check if HEIC/HEIF - need to convert before serving
		if file_path.suffix.lower() in ['.heic', '.heif']:
			# Read and convert to JPEG
			heic_content = file_path.read_bytes()
			from .pipeline import _convert_heic_to_jpeg
			jpeg_content = _convert_heic_to_jpeg(heic_content)
			
			# Return JPEG directly
			from fastapi.responses import Response
			return Response(
				content=jpeg_content,
				media_type="image/jpeg",
				headers={
					"Content-Disposition": f'inline; filename="{file_path.stem}.jpg"'
				}
			)
		
		# For other file types, serve as-is
		import mimetypes
		media_type, _ = mimetypes.guess_type(str(file_path))
		if not media_type:
			media_type = 'application/octet-stream'
		
		return FileResponse(
			path=str(file_path),
			filename=file_path.name,
			media_type=media_type
		)
	
	else:
		return {"error": "Invalid path format"}


# ----------------------------
# Database setup and endpoints
# ----------------------------

class SaveInvoice(BaseModel):
    supplier_id: Optional[int] = None
    supplier_name: Optional[str] = None
    invoice_number: str
    invoice_date: date
    currency: str
    subtotal: float
    vat_amount: float
    total: float
    expense_account_id: Optional[int] = None
    deductible_pct: Optional[float] = None
    doc_name: Optional[str] = None
    doc_full_path: Optional[str] = None
    document_type: Optional[str] = Field(default="invoice")
    status: Optional[str] = Field(default="pending")
    ocr_confidence: Optional[float] = None
    ocr_language: Optional[str] = None
    ocr_metadata: Optional[Dict[str, Any]] = None
    needs_review: Optional[bool] = None
    due_date: Optional[date] = None
    payment_terms: Optional[str] = None


class BatchSaveRequest(BaseModel):
    customer_id: int
    invoices: List[SaveInvoice]


class BatchSaveResult(BaseModel):
    index: int
    invoice_number: str
    inserted_id: Optional[int] = None
    supplier_id: Optional[int] = None
    supplier_created: Optional[bool] = None
    is_update: bool = False
    conflict: bool = False
    error: Optional[str] = None


@app.on_event("startup")
async def startup_db():
    # Engine is created lazily via get_engine()
    pass


@app.on_event("shutdown")
async def shutdown_db():
    await close_engine()


async def _ensure_supplier(conn, customer_id: int, supplier_id: Optional[int], supplier_name: Optional[str]) -> tuple[int, bool]:
    """Ensure supplier exists, update if needed, or create new one."""
    from sqlalchemy import func
    
    if supplier_id is not None:
        # If supplier_id exists and name is provided, update the supplier name
        if supplier_name:
            stmt = (
                update(suppliers)
                .where(suppliers.c.id == supplier_id, suppliers.c.customer_id == customer_id)
                .values(name=supplier_name, updated_at=func.now())
            )
            await conn.execute(stmt)
        return supplier_id, False
    
    if not supplier_name:
        raise ValueError("supplier_id or supplier_name is required")
    
    # Try match by OCR identification first, then by name
    stmt = select(suppliers.c.id).where(
        suppliers.c.customer_id == customer_id,
        (suppliers.c.ocr_supplier_identification == supplier_name) | (suppliers.c.name == supplier_name)
    )
    result = await conn.execute(stmt)
    row = result.first()
    if row:
        return int(row.id), False
    
    # Create new supplier
    stmt = (
        insert(suppliers)
        .values(
            customer_id=customer_id,
            name=supplier_name,
            ocr_supplier_identification=supplier_name,
            active=True
        )
        .returning(suppliers.c.id)
    )
    result = await conn.execute(stmt)
    row = result.first()
    return int(row.id), True


@app.post("/invoices/batch")
async def save_invoices_batch(payload: BatchSaveRequest) -> Dict[str, Any]:
    from sqlalchemy import func
    from .pipeline import _move_to_processed
    
    engine = get_engine()
    results: List[BatchSaveResult] = []
    
    # Prepare processed directory
    input_dir = Path(settings.upload_dir)
    processed_dir = input_dir.parent / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)
    
    async with engine.begin() as conn:  # Automatically commits or rolls back
        for idx, inv in enumerate(payload.invoices):
            try:
                sup_id, created = await _ensure_supplier(conn, payload.customer_id, inv.supplier_id, inv.supplier_name)
                
                # Check if invoice already exists (to determine INSERT vs UPDATE)
                check_stmt = select(invoices.c.id).where(
                    invoices.c.customer_id == payload.customer_id,
                    invoices.c.supplier_id == sup_id,
                    invoices.c.invoice_number == inv.invoice_number
                )
                check_result = await conn.execute(check_stmt)
                existing_row = check_result.first()
                is_update = existing_row is not None
                
                # Handle file moving if doc_full_path is provided
                final_doc_path = inv.doc_full_path
                if inv.doc_full_path:
                    # Extract file path from various formats
                    file_path_str = inv.doc_full_path
                    if file_path_str.startswith("file://"):
                        file_path_str = file_path_str[7:]
                    
                    file_path = Path(file_path_str)
                    
                    # Check if file is in input directory and needs to be moved
                    if file_path.exists() and 'input' in file_path.parts:
                        # Move file to processed directory
                        new_path = _move_to_processed(inv.doc_full_path)
                        if new_path != inv.doc_full_path:
                            final_doc_path = new_path
                            print(f"[INFO] Moved file to processed: {file_path.name}")
                    elif 'processed' in file_path.parts:
                        # File already in processed, keep as is
                        final_doc_path = inv.doc_full_path
                    else:
                        # File path doesn't exist or not in expected location
                        print(f"[WARN] File not found or not in expected location: {file_path_str}")
                
                # Build upsert statement with updated path
                stmt = pg_insert(invoices).values(
                    customer_id=payload.customer_id,
                    supplier_id=sup_id,
                    invoice_number=inv.invoice_number,
                    invoice_date=inv.invoice_date,
                    due_date=inv.due_date,
                    currency=inv.currency,
                    subtotal=inv.subtotal,
                    vat_amount=inv.vat_amount,
                    total=inv.total,
                    expense_account_id=inv.expense_account_id,
                    deductible_pct=inv.deductible_pct,
                    doc_name=inv.doc_name,
                    doc_full_path=final_doc_path,  # Use updated path
                    document_type=inv.document_type,
                    status=inv.status,
                    ocr_confidence=inv.ocr_confidence,
                    ocr_language=inv.ocr_language,
                    ocr_metadata=inv.ocr_metadata,
                    needs_review=inv.needs_review,
                    payment_terms=inv.payment_terms,
                )
                
                # On conflict, update all fields
                stmt = stmt.on_conflict_do_update(
                    index_elements=['customer_id', 'supplier_id', 'invoice_number'],
                    set_={
                        'invoice_date': stmt.excluded.invoice_date,
                        'due_date': stmt.excluded.due_date,
                        'currency': stmt.excluded.currency,
                        'subtotal': stmt.excluded.subtotal,
                        'vat_amount': stmt.excluded.vat_amount,
                        'total': stmt.excluded.total,
                        'expense_account_id': stmt.excluded.expense_account_id,
                        'deductible_pct': stmt.excluded.deductible_pct,
                        'doc_name': stmt.excluded.doc_name,
                        'doc_full_path': stmt.excluded.doc_full_path,
                        'document_type': stmt.excluded.document_type,
                        'status': stmt.excluded.status,
                        'ocr_confidence': stmt.excluded.ocr_confidence,
                        'ocr_language': stmt.excluded.ocr_language,
                        'ocr_metadata': stmt.excluded.ocr_metadata,
                        'needs_review': stmt.excluded.needs_review,
                        'payment_terms': stmt.excluded.payment_terms,
                        'updated_at': func.now(),
                    }
                ).returning(invoices.c.id)
                
                result = await conn.execute(stmt)
                row = result.first()
                
                if row and row.id is not None:
                    results.append(BatchSaveResult(
                        index=idx,
                        invoice_number=inv.invoice_number,
                        inserted_id=int(row.id),
                        supplier_id=sup_id,
                        supplier_created=created,
                        is_update=is_update
                    ))
                else:
                    results.append(BatchSaveResult(
                        index=idx,
                        invoice_number=inv.invoice_number,
                        inserted_id=None,
                        supplier_id=sup_id,
                        supplier_created=created,
                        is_update=is_update,
                        error="No ID returned from database"
                    ))
            except Exception as e:
                results.append(BatchSaveResult(
                    index=idx,
                    invoice_number=inv.invoice_number,
                    supplier_id=sup_id if 'sup_id' in locals() else None,
                    supplier_created=created if 'created' in locals() else None,
                    is_update=False,
                    error=f"{type(e).__name__}: {str(e)}"
                ))
    
    return {"results": [r.dict() for r in results]}


class ConflictCheckRequest(BaseModel):
    customer_id: int
    invoices: List[SaveInvoice]


class ConflictDetail(BaseModel):
    invoice_number: str
    type: str  # "db_constraint" or "filename_duplicate"
    message: str


class ConflictCheckResponse(BaseModel):
    has_conflicts: bool
    conflicts: List[ConflictDetail]


@app.post("/invoices/check-conflicts", response_model=ConflictCheckResponse)
async def check_invoice_conflicts(payload: ConflictCheckRequest) -> ConflictCheckResponse:
    """
    Check for conflicts before saving invoices.
    Checks:
    1. DB unique constraint (customer_id, supplier_id, invoice_number)
    2. Filename duplicates in processed folder
    """
    engine = get_engine()
    conflicts: List[ConflictDetail] = []
    
    # Prepare processed directory path
    input_dir = Path(settings.upload_dir)
    processed_dir = input_dir.parent / "processed"
    
    async with engine.connect() as conn:
        for inv in payload.invoices:
            try:
                # Get supplier_id
                sup_id = inv.supplier_id
                if not sup_id and inv.supplier_name:
                    # Try to find existing supplier
                    stmt = select(suppliers.c.id).where(
                        suppliers.c.customer_id == payload.customer_id,
                        (suppliers.c.ocr_supplier_identification == inv.supplier_name) | (suppliers.c.name == inv.supplier_name)
                    )
                    result = await conn.execute(stmt)
                    row = result.first()
                    if row:
                        sup_id = int(row.id)
                
                if not sup_id:
                    # Will create new supplier, no conflict check needed for DB
                    continue
                
                # Check DB constraint (only for invoices that will be INSERTs, not UPDATEs)
                check_stmt = select(invoices.c.id).where(
                    invoices.c.customer_id == payload.customer_id,
                    invoices.c.supplier_id == sup_id,
                    invoices.c.invoice_number == inv.invoice_number
                )
                check_result = await conn.execute(check_stmt)
                existing_row = check_result.first()
                
                # If it exists, it's an UPDATE, not a conflict
                if not existing_row:
                    # Check for filename conflicts in processed folder
                    if inv.doc_name and processed_dir.exists():
                        processed_files = list(processed_dir.glob("*"))
                        for processed_file in processed_files:
                            if processed_file.name == inv.doc_name or processed_file.name.endswith(f"_{inv.doc_name}"):
                                conflicts.append(ConflictDetail(
                                    invoice_number=inv.invoice_number,
                                    type="filename_duplicate",
                                    message=f"File '{inv.doc_name}' already exists in processed folder"
                                ))
                                break
                
            except Exception as e:
                print(f"[WARN] Error checking conflicts for invoice {inv.invoice_number}: {str(e)}")
                # Continue checking other invoices
                continue
    
    return ConflictCheckResponse(
        has_conflicts=len(conflicts) > 0,
        conflicts=conflicts
    )


@app.get("/customers")
async def get_customers() -> List[Dict[str, Any]]:
    """Fetch all active customers for dropdown selection."""
    engine = get_engine()
    
    async with engine.connect() as conn:
        stmt = select(
            customers.c.id,
            customers.c.name
        ).where(customers.c.active == True).order_by(customers.c.name)
        
        result = await conn.execute(stmt)
        rows = result.fetchall()
        
        return [{"id": row.id, "name": row.name} for row in rows]


@app.get("/invoices/report")
async def get_invoices_report(
    customer_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None  # Comma-separated status values
) -> List[Dict[str, Any]]:
    """
    Fetch invoices with filters for reporting.
    Query params:
    - customer_id (required)
    - start_date (optional, YYYY-MM-DD)
    - end_date (optional, YYYY-MM-DD)
    - status (optional, comma-separated: pending,approved,exported,rejected)
    """
    engine = get_engine()
    
    async with engine.connect() as conn:
        # Join invoices with suppliers to get supplier name
        j = join(invoices, suppliers, invoices.c.supplier_id == suppliers.c.id)
        
        stmt = select(
            invoices.c.id,
            invoices.c.invoice_date,
            invoices.c.subtotal,
            invoices.c.vat_amount,
            invoices.c.total,
            suppliers.c.name.label('supplier_name'),
            invoices.c.doc_name,
            invoices.c.status,
            invoices.c.invoice_number,
            invoices.c.currency
        ).select_from(j).where(
            invoices.c.customer_id == customer_id
        )
        
        # Apply date filters
        if start_date:
            from datetime import datetime
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            stmt = stmt.where(invoices.c.invoice_date >= start)
        
        if end_date:
            from datetime import datetime
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            stmt = stmt.where(invoices.c.invoice_date <= end)
        
        # Apply status filter
        if status:
            status_list = [s.strip() for s in status.split(',') if s.strip()]
            if status_list:
                stmt = stmt.where(invoices.c.status.in_(status_list))
        
        # Order by date descending
        stmt = stmt.order_by(invoices.c.invoice_date.desc())
        
        result = await conn.execute(stmt)
        rows = result.fetchall()
        
        return [
            {
                "id": row.id,
                "invoice_date": row.invoice_date.isoformat() if row.invoice_date else None,
                "subtotal": float(row.subtotal) if row.subtotal else 0.0,
                "vat_amount": float(row.vat_amount) if row.vat_amount else 0.0,
                "total": float(row.total) if row.total else 0.0,
                "supplier_name": row.supplier_name,
                "doc_name": row.doc_name,
                "status": row.status,
                "invoice_number": row.invoice_number,
                "currency": row.currency
            }
            for row in rows
        ]


class ExportInvoicesRequest(BaseModel):
    customer_id: int
    invoice_ids: Optional[List[int]] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: Optional[str] = None


@app.post("/invoices/export")
async def export_invoices(payload: ExportInvoicesRequest):
    """
    Export invoices to Excel file.
    If invoice_ids provided, export only those records.
    Otherwise apply filters like /invoices/report.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    engine = get_engine()
    
    async with engine.connect() as conn:
        # Join invoices with suppliers to get supplier name
        j = join(invoices, suppliers, invoices.c.supplier_id == suppliers.c.id)
        
        stmt = select(
            invoices.c.invoice_date,
            invoices.c.subtotal,
            invoices.c.vat_amount,
            invoices.c.total,
            invoices.c.currency,
            suppliers.c.name.label('supplier_name'),
            invoices.c.doc_name,
            invoices.c.status
        ).select_from(j).where(
            invoices.c.customer_id == payload.customer_id
        )
        
        # If specific invoice IDs provided, use those
        if payload.invoice_ids:
            stmt = stmt.where(invoices.c.id.in_(payload.invoice_ids))
        else:
            # Apply date filters
            if payload.start_date:
                from datetime import datetime
                start = datetime.strptime(payload.start_date, '%Y-%m-%d').date()
                stmt = stmt.where(invoices.c.invoice_date >= start)
            
            if payload.end_date:
                from datetime import datetime
                end = datetime.strptime(payload.end_date, '%Y-%m-%d').date()
                stmt = stmt.where(invoices.c.invoice_date <= end)
            
            # Apply status filter
            if payload.status:
                status_list = [s.strip() for s in payload.status.split(',') if s.strip()]
                if status_list:
                    stmt = stmt.where(invoices.c.status.in_(status_list))
        
        # Order by date
        stmt = stmt.order_by(invoices.c.invoice_date.desc())
        
        result = await conn.execute(stmt)
        rows = result.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices Report"
        
        # Header row
        headers = ["Invoice Date", "Amount", "VAT", "Total Amount", "Currency", "Supplier Name", "Doc Name", "Status"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Data rows
        for row in rows:
            ws.append([
                row.invoice_date.isoformat() if row.invoice_date else "",
                float(row.subtotal) if row.subtotal else 0.0,
                float(row.vat_amount) if row.vat_amount else 0.0,
                float(row.total) if row.total else 0.0,
                row.currency or "",
                row.supplier_name or "",
                row.doc_name or "",
                row.status or ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as downloadable file
        from datetime import datetime
        filename = f"invoices_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )


